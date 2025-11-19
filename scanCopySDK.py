#!/usr/bin/env python3
"""
Scan Copy SDK - Interactive tool for copying scans and generating analysis
"""

import json
import os
import sys
import subprocess
import pandas as pd
import glob
from datetime import datetime
from getpass import getpass
import config

# Clipboard functionality removed as requested

class ScanCopySDK:
    def __init__(self):
        self.config_file = "config.json"
        self.config = {}
        self.source_scan_ids = []
        self.target_scan_ids = []
        self.target_store_id = None
        self.scan_mapping = {}
        self.run_folder = None
        self.custom_results_path = None
        self.checkpoint_prompt_shown = False

    def handle_checkpoint_resume(self):
        """Offer user the choice to resume from or restart without checkpoint."""
        if self.checkpoint_prompt_shown:
            return

        checkpoint_files = glob.glob("checkpoint_*.json")
        if not checkpoint_files:
            return

        latest_checkpoint = max(checkpoint_files, key=os.path.getmtime)
        print("\n" + "=" * 80)
        print("Checkpoint detected!")
        print(f"Latest checkpoint file: {latest_checkpoint}")
        print("You can resume from this checkpoint or restart from scratch.")

        while True:
            choice = input("Resume from checkpoint? (y = resume, n = restart): ").strip().lower()
            if choice in ("y", "yes"):
                print("✅ Resuming from existing checkpoint...\n")
                self.checkpoint_prompt_shown = True
                return
            if choice in ("n", "no"):
                print("🔁 Restarting from scratch. Removing checkpoint files...")
                for cp in checkpoint_files:
                    try:
                        os.remove(cp)
                        print(f"🗑️  Removed {cp}")
                    except Exception as e:
                        print(f"⚠️  Could not remove {cp}: {e}")
                self.checkpoint_prompt_shown = True
                return
            print("Please enter 'y' to resume or 'n' to restart.")
    
    def create_results_folder(self):
        """Create a timestamped folder for this run's results"""
        try:
            # Use custom path if provided, otherwise use default
            if self.custom_results_path:
                base_dir = self.custom_results_path
            else:
                current_dir = os.getcwd()
                base_dir = os.path.join(current_dir, "testResults")
            
            # Create base directory if it doesn't exist
            if not os.path.exists(base_dir):
                os.makedirs(base_dir, exist_ok=True)
                print(f"📁 Created results directory: {base_dir}")
            
            # Create timestamped subfolder for this run
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.run_folder = os.path.join(base_dir, f"run_{timestamp}")
            
            # Create the run folder with exist_ok=True to handle race conditions
            os.makedirs(self.run_folder, exist_ok=True)
            
            print(f"📁 Created run folder: {self.run_folder}")
            return True
            
        except PermissionError as e:
            print(f"❌ Permission denied creating folder: {e}")
            print("💡 Try running as administrator or check folder permissions")
            return False
        except OSError as e:
            print(f"❌ OS Error creating folder: {e}")
            print("💡 Check if the path is valid and you have write permissions")
            return False
        except Exception as e:
            print(f"❌ Unexpected error creating results folder: {e}")
            print("🔄 Trying fallback folder creation...")
            
            # Fallback: try to create a simple folder
            try:
                fallback_folder = os.path.join(current_dir, "testResults", "current_run")
                os.makedirs(fallback_folder, exist_ok=True)
                self.run_folder = fallback_folder
                print(f"📁 Created fallback run folder: {self.run_folder}")
                return True
            except Exception as fallback_error:
                print(f"❌ Fallback folder creation also failed: {fallback_error}")
                print("💡 Please check your permissions and try running as administrator")
                import traceback
                traceback.print_exc()
                return False
    
    def move_mapping_file_to_run_folder(self):
        """Move the scan mapping file from current directory to run folder"""
        try:
            # Look for scan mapping CSV files in current directory
            csv_files = [f for f in os.listdir('.') if f.startswith('scan_mapping') and f.endswith('.csv')]
            
            if csv_files:
                # Get the most recent one
                latest_csv = max(csv_files, key=os.path.getctime)
                source_path = os.path.abspath(latest_csv)
                dest_path = os.path.join(self.run_folder, latest_csv)
                
                # Ensure destination directory exists
                os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                
                # Move the file using shutil for better cross-platform support
                import shutil
                shutil.move(source_path, dest_path)
                print(f"📁 Moved mapping file to run folder: {latest_csv}")
            else:
                print("⚠️  No scan mapping file found to move")
                
        except PermissionError as e:
            print(f"⚠️  Permission denied moving file: {e}")
            print("💡 File might be in use or you need administrator privileges")
        except OSError as e:
            print(f"⚠️  OS Error moving file: {e}")
            print("💡 Check if source file exists and destination is writable")
        except Exception as e:
            print(f"⚠️  Error moving mapping file: {e}")
            import traceback
            traceback.print_exc()
        
    def step1_configuration(self):
        """Step 1: Collect configuration details from user or use existing config"""
        print("=" * 60)
        print("STEP 1: CONFIGURATION")
        print("=" * 60)
        
        try:
            # Try to load existing configuration first from config.json
            if os.path.exists('config.json'):
                with open('config.json', 'r') as f:
                    config_data = json.load(f)
                
                # Check if config has values
                has_config = (config_data.get('SOURCE_INSTANCE') and config_data.get('SOURCE_DB_PASSWORD') and 
                             config_data.get('SOURCE_USERNAME') and config_data.get('SOURCE_PASSWORD') and 
                             config_data.get('TARGET_INSTANCE') and config_data.get('TARGET_DB_PASSWORD') and 
                             config_data.get('TARGET_USERNAME') and config_data.get('TARGET_PASSWORD'))
                
                if has_config:
                    print("✅ Found existing configuration in config.json")
                    print(f"   Source Instance: {config_data.get('SOURCE_INSTANCE')}")
                    print(f"   Target Instance: {config_data.get('TARGET_INSTANCE')}")
                    print(f"   Source Username: {config_data.get('SOURCE_USERNAME')}")
                    print(f"   Target Username: {config_data.get('TARGET_USERNAME')}")
                    
                    use_existing = input("\nUse existing configuration? (y/n): ").strip().lower()
                    if use_existing in ['y', 'yes']:
                        # Use existing configuration
                        self.config = config_data
                        print("✅ Using existing configuration")
                        return True
            
            print("🔧 Please enter your database configuration details:")
            print()
            
            # Collect source database details
            print("📊 SOURCE DATABASE:")
            source_instance = input("Source Instance (e.g., albt): ").strip()
            if not source_instance:
                print("❌ Source instance is required")
                return False
            
            source_db_password = getpass("Source DB Password: ").strip()
            if not source_db_password:
                print("❌ Source DB password is required")
                return False
            
            source_username = input("Source Username: ").strip()
            if not source_username:
                print("❌ Source username is required")
                return False
            
            source_password = getpass("Source Password: ").strip()
            if not source_password:
                print("❌ Source password is required")
                return False
            
            print()
            
            # Collect target database details
            print("📊 TARGET DATABASE:")
            target_instance = input("Target Instance (e.g., stgalbt): ").strip()
            if not target_instance:
                print("❌ Target instance is required")
                return False
            
            target_db_password = getpass("Target DB Password: ").strip()
            if not target_db_password:
                print("❌ Target DB password is required")
                return False
            
            target_username = input("Target Username: ").strip()
            if not target_username:
                print("❌ Target username is required")
                return False
            
            target_password = getpass("Target Password: ").strip()
            if not target_password:
                print("❌ Target password is required")
                return False
            
            # Store configuration
            self.config = {
                'SOURCE_INSTANCE': source_instance,
                'SOURCE_DB_PASSWORD': source_db_password,
                'SOURCE_USERNAME': source_username,
                'SOURCE_PASSWORD': source_password,
                'TARGET_INSTANCE': target_instance,
                'TARGET_DB_PASSWORD': target_db_password,
                'TARGET_USERNAME': target_username,
                'TARGET_PASSWORD': target_password
            }
            
            # Save configuration to config.json
            with open(self.config_file, 'w') as f:
                json.dump(self.config, f, indent=2)
            
            print()
            print("✅ Configuration saved successfully!")
            print(f"   Source Instance: {self.config['SOURCE_INSTANCE']}")
            print(f"   Target Instance: {self.config['TARGET_INSTANCE']}")
            print(f"   Source DB Password: {'*' * len(self.config['SOURCE_DB_PASSWORD'])}")
            print(f"   Target DB Password: {'*' * len(self.config['TARGET_DB_PASSWORD'])}")
            print(f"   Source Username: {self.config['SOURCE_USERNAME']}")
            print(f"   Target Username: {self.config['TARGET_USERNAME']}")
            print(f"   Source Password: {'*' * len(self.config['SOURCE_PASSWORD'])}")
            print(f"   Target Password: {'*' * len(self.config['TARGET_PASSWORD'])}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error collecting configuration: {e}")
            return False
    
    def step1_5_get_results_path(self):
        """Step 1.5: Get custom results path from user"""
        print("\n" + "=" * 60)
        print("STEP 1.5: RESULTS PATH")
        print("=" * 60)
        
        print("📁 Choose where to save the test results:")
        print("1. Use default path (./testResults/)")
        print("2. Specify custom path")
        
        while True:
            choice = input("Enter your choice (1-2): ").strip()
            
            if choice == '1':
                self.custom_results_path = None
                print("✅ Using default results path: ./testResults/")
                return True
            elif choice == '2':
                custom_path = input("Enter custom path for results: ").strip()
                if not custom_path:
                    print("❌ Please enter a valid path")
                    continue
                
                # Validate path
                try:
                    # Test if we can create the directory
                    os.makedirs(custom_path, exist_ok=True)
                    self.custom_results_path = os.path.abspath(custom_path)
                    print(f"✅ Custom results path set: {self.custom_results_path}")
                    return True
                except Exception as e:
                    print(f"❌ Invalid path: {e}")
                    print("💡 Please enter a valid, writable directory path")
                    continue
            else:
                print("❌ Invalid choice. Please enter 1 or 2.")
    
    def step2_get_source_scan_ids(self):
        """Step 2: Get source scan IDs from user or use existing config"""
        print("\n" + "=" * 60)
        print("STEP 2: SOURCE SCAN IDs")
        print("=" * 60)
        
        # Check if config has scan IDs
        import config
        if config.SCAN_IDS_FOR_COPYING:
            print(f"✅ Found existing scan IDs in config.py: {list(config.SCAN_IDS_FOR_COPYING)}")
            use_existing = input("Use existing scan IDs? (y/n): ").strip().lower()
            if use_existing in ['y', 'yes']:
                self.source_scan_ids = list(config.SCAN_IDS_FOR_COPYING)
                print(f"✅ Using existing scan IDs: {self.source_scan_ids}")
                # Create initial mapping file with source scan IDs
                self.create_initial_mapping_file()
                return True
        
        # Ask user to choose input method
        print("\nChoose how to get source scan IDs:")
        print("1. Enter scan IDs manually (comma-separated)")
        print("2. Get scan IDs automatically from database (by date and store ID)")
        print("3. Skip source scan IDs (create empty mapping file)")
        
        while True:
            choice = input("Enter choice (1, 2, or 3): ").strip()
            if choice == '1':
                # Manual input
                while True:
                    scan_input = input("Enter source scan IDs (comma-separated): ").strip()
                    if not scan_input:
                        print("❌ Please enter at least one scan ID")
                        continue
                    
                    try:
                        self.source_scan_ids = self._parse_scan_ids(scan_input)
                        print(f"✅ Source scan IDs: {self.source_scan_ids}")
                        # Create initial mapping file with source scan IDs
                        self.create_initial_mapping_file()
                        return True
                    except ValueError as e:
                        print(f"❌ {e}")
                        
            elif choice == '2':
                # Automatic retrieval from database
                try:
                    self.source_scan_ids = self.get_scan_ids_from_database()
                    if self.source_scan_ids:
                        print(f"✅ Retrieved {len(self.source_scan_ids)} scan IDs from database")
                        print(f"✅ Source scan IDs: {self.source_scan_ids}")
                        # Create initial mapping file with source scan IDs
                        self.create_initial_mapping_file()
                        return True
                    else:
                        print("❌ No scan IDs found for the specified criteria")
                        continue
                except Exception as e:
                    print(f"❌ Error retrieving scan IDs from database: {e}")
                    continue
                    
            elif choice == '3':
                # Skip source scan IDs - create empty mapping file
                self.source_scan_ids = []
                print("✅ Skipping source scan IDs - will create empty mapping file")
                # Create initial mapping file (empty)
                self.create_initial_mapping_file()
                return True
            else:
                print("❌ Invalid choice. Please enter 1, 2, or 3.")
    
    def get_scan_ids_from_database(self):
        """Get scan IDs from realograms_implementation_scan table based on date and store_id"""
        try:
            import psycopg
            from psycopg.rows import dict_row
            from datetime import datetime
            
            print("\n📊 Database Query for Scan IDs")
            print("=" * 40)
            
            # Get date from user
            while True:
                date_input = input("Enter scan date (YYYY-MM-DD format): ").strip()
                if not date_input:
                    print("❌ Please enter a date")
                    continue
                
                try:
                    # Validate date format
                    scan_date = datetime.strptime(date_input, '%Y-%m-%d').date()
                    break
                except ValueError:
                    print("❌ Invalid date format. Please use YYYY-MM-DD format (e.g., 2024-01-15)")
                    continue
            
            # Get store ID from user
            while True:
                store_input = input("Enter store ID: ").strip()
                if not store_input:
                    print("❌ Please enter a store ID")
                    continue
                
                try:
                    store_id = int(store_input)
                    break
                except ValueError:
                    print("❌ Please enter a valid number for store ID")
                    continue
            
            print(f"\n🔍 Querying database for scans on {scan_date} in store {store_id}...")
            
            # Connect to source database
            with psycopg.connect(
                user='proxyuser',
                password=self.config['SOURCE_DB_PASSWORD'],
                host=f"{self.config['SOURCE_INSTANCE']}-maint.rebotics.net",
                port='5432',
                dbname=self.config['SOURCE_INSTANCE'],
                row_factory=dict_row
            ) as connection:
                with connection.cursor() as cursor:
                    # Query to get scan IDs for the specified date and store
                    query = """
                        SELECT id, created_at, store_id
                        FROM realograms_implementation_scan 
                        WHERE DATE(created_at) = %s 
                        AND store_id = %s
                        ORDER BY created_at DESC;
                    """
                    
                    cursor.execute(query, (scan_date, store_id))
                    results = cursor.fetchall()
                    
                    if results:
                        scan_ids = [row['id'] for row in results]
                        print(f"✅ Found {len(scan_ids)} scans:")
                        for i, row in enumerate(results[:10]):  # Show first 10
                            print(f"   {i+1}. Scan ID: {row['id']}, Created: {row['created_at']}")
                        if len(results) > 10:
                            print(f"   ... and {len(results) - 10} more scans")
                        
                        return scan_ids
                    else:
                        print(f"❌ No scans found for date {scan_date} in store {store_id}")
                        return []
                        
        except Exception as e:
            print(f"❌ Database error: {e}")
            return []
    
    def create_initial_mapping_file(self):
        """Create initial mapping file with source scan IDs (target IDs will be empty initially)"""
        try:
            import csv
            from datetime import datetime
            
            # Create initial mapping file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            mapping_filename = f"initial_scan_mapping_{timestamp}.csv"
            
            with open(mapping_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Source_Scan_ID', 'Target_Scan_ID'])
                
                # Write source scan IDs with empty target IDs (even if empty list)
                if self.source_scan_ids:
                    for scan_id in self.source_scan_ids:
                        writer.writerow([scan_id, ''])
                else:
                    # Add a placeholder row if no source scan IDs
                    writer.writerow(['', ''])
            
            print(f"📄 Created initial mapping file: {mapping_filename}")
            if self.source_scan_ids:
                print(f"   Source scan IDs: {self.source_scan_ids}")
            else:
                print(f"   Source scan IDs: (empty - ready for manual entry)")
            print(f"   Target scan IDs: (to be filled after copy operation)")
            
            # Move to run folder if it exists
            if hasattr(self, 'run_folder') and self.run_folder:
                try:
                    import shutil
                    dest_path = os.path.join(self.run_folder, mapping_filename)
                    shutil.move(mapping_filename, dest_path)
                    print(f"📁 Moved to run folder: {dest_path}")
                except Exception as e:
                    print(f"⚠️  Could not move to run folder: {e}")
            
        except Exception as e:
            print(f"⚠️  Error creating initial mapping file: {e}")
    
    def _parse_scan_ids(self, text):
        """Parse scan IDs from text"""
        if not text:
            raise ValueError("No scan IDs provided")
        
        # Parse scan IDs
        scan_ids = [int(x.strip()) for x in text.split(',') if x.strip()]
        
        if not scan_ids:
            raise ValueError("No valid scan IDs found")
        
        # Validate scan IDs
        if not all(isinstance(scan_id, int) and scan_id > 0 for scan_id in scan_ids):
            raise ValueError("All scan IDs must be positive integers")
        
        return scan_ids
    
    def step3_get_target_store(self):
        """Step 3: Get target store ID from user or use existing config"""
        print("\n" + "=" * 60)
        print("STEP 3: TARGET STORE")
        print("=" * 60)
        
        # Check if config has target store ID
        import config
        if config.TARGET_STORE_ID is not None:
            print(f"✅ Found existing target store ID in config: {config.TARGET_STORE_ID}")
            use_existing = input("Use existing target store ID? (y/n): ").strip().lower()
            if use_existing in ['y', 'yes']:
                self.target_store_id = config.TARGET_STORE_ID
                print(f"✅ Using existing target store ID: {self.target_store_id}")
                return True
        
        while True:
            store_input = input("Enter target store ID: ").strip()
            if not store_input:
                print("❌ Please enter a store ID")
                continue
            
            try:
                self.target_store_id = int(store_input)
                print(f"✅ Target store ID: {self.target_store_id}")
                return True
            except ValueError:
                print("❌ Invalid input. Please enter a number only.")
    
    def step4_run_copy_script(self):
        """Step 4: Run copy script and generate CSV"""
        print("\n" + "=" * 60)
        print("STEP 4: COPY SCANS")
        print("=" * 60)
        
        # Ask user to choose script
        while True:
            choice = input("Choose an option:\n1. copyScans.py\n2. copyScanUpdated.py\n3. Skip copy operation (mapping file already created)\nEnter choice (1, 2, or 3): ").strip()
            if choice in ['1', '2', '3']:
                break
            print("❌ Please enter 1, 2, or 3")
        
        if choice == '3':
            print("✅ Skipping copy operation")
            print("📄 Using existing initial mapping file")
            return True
        
        script_name = "copyScans.py" if choice == '1' else "copyScanUpdated.py"
        print(f"✅ Selected script: {script_name}")
        
        # Update config with current values
        self.update_config_with_current_values()
        
        # Prompt user about checkpoint resume/restart if applicable
        self.handle_checkpoint_resume()

        # Run the selected script
        try:
            print(f"\n🔄 Running {script_name}...")
            print("=" * 80)
            
            # Use Popen to stream output in real-time instead of buffering
            process = subprocess.Popen(
                [sys.executable, script_name],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,  # Line buffered
                universal_newlines=True
            )
            
            # Stream output in real-time with timeout handling
            output_lines = []
            import threading
            import time
            
            def read_output():
                for line in process.stdout:
                    output_lines.append(line)
                    print(line, end='', flush=True)
            
            # Start output reading thread
            output_thread = threading.Thread(target=read_output, daemon=True)
            output_thread.start()
            
            # Wait for process to complete with increased timeout (30 minutes for large batches)
            start_time = time.time()
            timeout_seconds = 1800  # 30 minutes
            
            while process.poll() is None:
                if time.time() - start_time > timeout_seconds:
                    raise subprocess.TimeoutExpired(process.args, timeout_seconds)
                time.sleep(0.1)  # Check every 100ms
            
            # Wait for output thread to finish reading remaining output
            output_thread.join(timeout=5)
            
            if process.returncode == 0:
                print("=" * 80)
                print(f"✅ {script_name} completed successfully")
                
                # Move mapping file to run folder
                self.move_mapping_file_to_run_folder()
                
                # Look for scan mapping CSV file
                self.find_scan_mapping_csv()
                return True
            else:
                print("=" * 80)
                print(f"❌ {script_name} failed with return code {process.returncode}")
                return False
                
        except subprocess.TimeoutExpired:
            print("=" * 80)
            print(f"❌ {script_name} timed out after 30 minutes")
            print("💡 Note: Progress has been saved to checkpoint file. You can resume by running again.")
            if process:
                process.kill()
            return False
        except Exception as e:
            print("=" * 80)
            print(f"❌ Error running {script_name}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def step5_get_target_scan_ids(self):
        """Step 5: Get target scan IDs (auto-extract from mapping or user input)"""
        print("\n" + "=" * 60)
        print("STEP 5: TARGET SCAN IDs")
        print("=" * 60)
        
        # Try to auto-extract target scan IDs from mapping
        if self.scan_mapping:
            auto_extracted_ids = list(self.scan_mapping.values())
            print(f"🔄 Auto-extracted target scan IDs from mapping: {auto_extracted_ids}")
            
            use_auto = input("Use auto-extracted target scan IDs? (y/n): ").strip().lower()
            if use_auto in ['y', 'yes']:
                self.target_scan_ids = auto_extracted_ids
                print(f"✅ Using auto-extracted target scan IDs: {self.target_scan_ids}")
                return True
        
        # Ask user to choose input method
        print("\nChoose how to get target scan IDs:")
        print("1. Enter scan IDs manually (comma-separated)")
        print("2. Select from mapping CSV file")
        
        while True:
            choice = input("Enter choice (1 or 2): ").strip()
            if choice == '1':
                # Manual input
                while True:
                    scan_input = input("Enter target scan IDs (comma-separated): ").strip()
                    if not scan_input:
                        print("❌ Please enter at least one scan ID")
                        continue
                    
                    try:
                        self.target_scan_ids = self._parse_scan_ids(scan_input)
                        print(f"✅ Target scan IDs: {self.target_scan_ids}")
                        return True
                    except ValueError as e:
                        print(f"❌ {e}")
                        
            elif choice == '2':
                # Select from mapping CSV file
                try:
                    self.target_scan_ids = self.select_target_scan_ids_from_mapping()
                    if self.target_scan_ids:
                        print(f"✅ Selected target scan IDs: {self.target_scan_ids}")
                        return True
                    else:
                        print("❌ No target scan IDs selected")
                        continue
                except Exception as e:
                    print(f"❌ Error reading mapping file: {e}")
                    continue
            else:
                print("❌ Invalid choice. Please enter 1 or 2.")
    
    def select_target_scan_ids_from_mapping(self):
        """Select target scan IDs from the mapping CSV file"""
        try:
            import csv
            import glob
            
            # Find mapping CSV files in the run folder
            mapping_files = []
            if hasattr(self, 'run_folder') and self.run_folder:
                # Look for updated mapping files first (with target scan IDs)
                updated_pattern = os.path.join(self.run_folder, "scan_mapping_updated_*.csv")
                mapping_files.extend(glob.glob(updated_pattern))
                
                # Fallback to initial mapping files if no updated ones found
                if not mapping_files:
                    initial_pattern = os.path.join(self.run_folder, "initial_scan_mapping_*.csv")
                    mapping_files.extend(glob.glob(initial_pattern))
            
            # Also look in current directory as fallback
            if not mapping_files:
                mapping_files.extend(glob.glob("scan_mapping_updated_*.csv"))
                if not mapping_files:
                    mapping_files.extend(glob.glob("initial_scan_mapping_*.csv"))
            
            if not mapping_files:
                print("❌ No mapping CSV files found")
                return []
            
            # Prefer updated mapping files over initial ones, then use most recent
            updated_files = [f for f in mapping_files if "scan_mapping_updated_" in f]
            if updated_files:
                mapping_file = max(updated_files, key=os.path.getctime)
                print(f"📄 Using updated mapping file: {os.path.basename(mapping_file)}")
            else:
                mapping_file = max(mapping_files, key=os.path.getctime)
                print(f"📄 Using initial mapping file: {os.path.basename(mapping_file)}")
            
            # Read the mapping file
            with open(mapping_file, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                rows = list(reader)
            
            if not rows:
                print("❌ Mapping file is empty")
                return []
            
            # Display available target scan IDs
            print("\n📋 Available target scan IDs in mapping file:")
            print("=" * 50)
            target_ids = []
            for i, row in enumerate(rows, 1):
                source_id = row.get('Source_Scan_ID', '').strip()
                target_id = row.get('Target_Scan_ID', '').strip()
                
                if target_id:  # Only show rows with target scan IDs
                    print(f"{i:2d}. Source: {source_id} → Target: {target_id}")
                    target_ids.append(int(target_id))
                elif source_id:  # Show rows with source but no target
                    print(f"{i:2d}. Source: {source_id} → Target: (empty)")
            
            if not target_ids:
                print("❌ No target scan IDs found in mapping file")
                print("💡 You may need to run the copy operation first to populate target IDs")
                return []
            
            # Let user select which target IDs to use
            print(f"\nFound {len(target_ids)} target scan IDs")
            print("Choose an option:")
            print("1. Use all target scan IDs")
            print("2. Select specific target scan IDs")
            
            while True:
                choice = input("Enter choice (1 or 2): ").strip()
                if choice == '1':
                    return target_ids
                elif choice == '2':
                    # Let user select specific IDs
                    while True:
                        selection = input(f"Enter target scan IDs to use (comma-separated, 1-{len(target_ids)}): ").strip()
                        if not selection:
                            print("❌ Please enter at least one selection")
                            continue
                        
                        try:
                            # Parse selection (could be indices or actual IDs)
                            selected_values = [x.strip() for x in selection.split(',')]
                            selected_ids = []
                            
                            for value in selected_values:
                                if value.isdigit():
                                    # Check if it's an index (1-based)
                                    idx = int(value) - 1
                                    if 0 <= idx < len(target_ids):
                                        selected_ids.append(target_ids[idx])
                                    else:
                                        print(f"❌ Invalid index: {value}")
                                        break
                                else:
                                    # Assume it's an actual scan ID
                                    try:
                                        scan_id = int(value)
                                        if scan_id in target_ids:
                                            selected_ids.append(scan_id)
                                        else:
                                            print(f"❌ Target scan ID {scan_id} not found in mapping")
                                            break
                                    except ValueError:
                                        print(f"❌ Invalid scan ID: {value}")
                                        break
                            else:
                                # All selections were valid
                                return selected_ids
                                
                        except ValueError as e:
                            print(f"❌ {e}")
                else:
                    print("❌ Invalid choice. Please enter 1 or 2.")
                    
        except Exception as e:
            print(f"❌ Error reading mapping file: {e}")
            return []
    
    def update_config_with_current_values(self):
        """Update config.py with current values"""
        try:
            # Read current config.py
            with open('config.py', 'r') as f:
                content = f.read()
            
            # Update database credentials
            content = self.update_config_value(content, 'SOURCE_INSTANCE', f"'{self.config['SOURCE_INSTANCE']}'")
            content = self.update_config_value(content, 'SOURCE_DB_PASSWORD', f"'{self.config['SOURCE_DB_PASSWORD']}'")
            content = self.update_config_value(content, 'SOURCE_USERNAME', f"'{self.config['SOURCE_USERNAME']}'")
            content = self.update_config_value(content, 'SOURCE_PASSWORD', f"'{self.config['SOURCE_PASSWORD']}'")
            content = self.update_config_value(content, 'TARGET_INSTANCE', f"'{self.config['TARGET_INSTANCE']}'")
            content = self.update_config_value(content, 'TARGET_DB_PASSWORD', f"'{self.config['TARGET_DB_PASSWORD']}'")
            content = self.update_config_value(content, 'TARGET_USERNAME', f"'{self.config['TARGET_USERNAME']}'")
            content = self.update_config_value(content, 'TARGET_PASSWORD', f"'{self.config['TARGET_PASSWORD']}'")
            
            # Update scan IDs
            content = self.update_config_value(content, 'SCAN_IDS_FOR_COPYING', f"({', '.join(map(str, self.source_scan_ids))},)")
            
            # Update target store ID
            content = self.update_config_value(content, 'TARGET_STORE_ID', str(self.target_store_id))
            
            # Write updated config.py
            with open('config.py', 'w') as f:
                f.write(content)
            
            print("✅ Updated config.py with current values")
            
        except Exception as e:
            print(f"❌ Error updating config.py: {e}")
    
    def update_config_value(self, content, key, value):
        """Update a specific value in config.py content"""
        import re
        
        if key == 'SCAN_IDS_FOR_COPYING':
            # Special handling for tuple values
            pattern = rf"^{key}\s*=\s*\([^)]*\)"
            replacement = f"{key} = {value}"
            return re.sub(pattern, replacement, content, flags=re.MULTILINE)
        else:
            # Regular handling for other values
            pattern = rf"^{key}\s*=\s*.*$"
            replacement = f"{key} = {value}"
            return re.sub(pattern, replacement, content, flags=re.MULTILINE)
    
    def find_scan_mapping_csv(self):
        """Find and load scan mapping CSV file"""
        try:
            # Look for scan mapping CSV files in the run folder
            csv_files = [f for f in os.listdir(self.run_folder) if f.startswith('scan_mapping') and f.endswith('.csv')]
            
            if csv_files:
                # Get the most recent one by creation time
                latest_csv = max(csv_files, key=lambda f: os.path.getctime(os.path.join(self.run_folder, f)))
                print(f"✅ Found scan mapping file: {latest_csv}")
                
                # Load the mapping
                df = pd.read_csv(os.path.join(self.run_folder, latest_csv))
                self.scan_mapping = dict(zip(df['Source_Scan_ID'], df['Target_Scan_ID']))
                print(f"✅ Loaded {len(self.scan_mapping)} scan mappings")
                
                # Show mapping
                print("\nScan ID Mappings:")
                for source_id, target_id in self.scan_mapping.items():
                    print(f"  Source: {source_id} -> Target: {target_id}")
            else:
                print("⚠️  No scan mapping CSV file found")
                
        except Exception as e:
            print(f"❌ Error loading scan mapping: {e}")
    
    def generate_analysis_csv(self):
        """Generate analysis CSV with source and target scan data"""
        print("\n" + "=" * 60)
        print("GENERATING ANALYSIS CSV")
        print("=" * 60)
        
        try:
            # Import the scan data analysis function
            from scanDataAnalysis import get_scan_data, check_additional_sections, process_scan_data, create_detailed_csv_report
            
            # Get source data
            print("🔄 Getting source scan data...")
            source_data = get_scan_data(
                self.config['SOURCE_INSTANCE'], 
                self.config['SOURCE_DB_PASSWORD'], 
                tuple(self.source_scan_ids)
            )
            
            # Get additional sections data
            additional_sections_data = check_additional_sections(
                self.config['SOURCE_INSTANCE'], 
                self.config['SOURCE_DB_PASSWORD'], 
                tuple(self.source_scan_ids)
            )
            
            # Process source data
            processed_source_data = process_scan_data(source_data, additional_sections_data)
            
            # Get target data if target scan IDs are provided
            processed_target_data = []
            if self.target_scan_ids:
                print("🔄 Getting target scan data...")
                target_data = get_scan_data(
                    self.config['TARGET_INSTANCE'], 
                    self.config['TARGET_DB_PASSWORD'], 
                    tuple(self.target_scan_ids)
                )
                
                target_additional_sections_data = check_additional_sections(
                    self.config['TARGET_INSTANCE'], 
                    self.config['TARGET_DB_PASSWORD'], 
                    tuple(self.target_scan_ids)
                )
                
                processed_target_data = process_scan_data(target_data, target_additional_sections_data)
            
            # Create source CSV with detailed information
            source_csv = create_detailed_csv_report(processed_source_data, "source_scandetails", self.run_folder)
            print(f"✅ Source scan details CSV: {source_csv}")
            
            # Create target CSV if target data exists
            target_csv = None
            if processed_target_data:
                target_csv = create_detailed_csv_report(processed_target_data, "target_scandetails", self.run_folder)
                print(f"✅ Target scan details CSV: {target_csv}")
            
            # Create analysis CSV with comments and highlighting
            if processed_target_data and source_csv and target_csv:
                # Use the actual file paths returned by create_detailed_csv_report
                try:
                    # Read the CSV files using the actual paths
                    source_df = pd.read_csv(source_csv)
                    target_df = pd.read_csv(target_csv)
                    
                    # Create analysis CSV with comments
                    analysis_csv = self.create_analysis_csv_with_comments_from_dataframes(source_df, target_df)
                    print(f"✅ Analysis CSV with comments: {analysis_csv}")
                except Exception as e:
                    print(f"⚠️  Error reading CSV files: {e}")
                    print(f"   Source CSV: {source_csv}")
                    print(f"   Target CSV: {target_csv}")
            else:
                print("⚠️  Cannot create analysis CSV - missing source or target data")
            
            return True
            
        except Exception as e:
            print(f"❌ Error generating analysis CSV: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def create_analysis_csv_with_comments_from_dataframes(self, source_df, target_df):
        """Create analysis CSV with comments and highlighting from DataFrames"""
        try:
            # Create analysis data
            analysis_data = []
            
            # Use scan mapping to match source and target scans
            print(f"📊 Scan mapping available: {self.scan_mapping}")
            print(f"📊 Source scans: {source_df['scan_id'].tolist()}")
            print(f"📊 Target scans: {target_df['scan_id'].tolist()}")
            
            for _, source_row in source_df.iterrows():
                source_scan_id = source_row['scan_id']
                
                # Find corresponding target scan ID from mapping
                target_scan_id = None
                if self.scan_mapping and source_scan_id in self.scan_mapping:
                    target_scan_id = self.scan_mapping[source_scan_id]
                    print(f"🔍 Looking for target scan ID: {target_scan_id} (mapped from source {source_scan_id})")
                
                # Find target scan by target scan ID
                target_matches = target_df[target_df['scan_id'] == target_scan_id] if target_scan_id else pd.DataFrame()
                
                if len(target_matches) > 0:
                    target_row = target_matches.iloc[0]
                    print(f"✅ Matched source scan {source_scan_id} with target scan {target_row['scan_id']}")
                else:
                    # Fallback: try to match by index if no mapping found
                    if len(target_df) > 0:
                        target_row = target_df.iloc[0]  # Use first target scan as fallback
                        print(f"⚠️  No mapping found for source scan {source_scan_id}, using fallback target scan {target_row['scan_id']}")
                    else:
                        print(f"❌ No target data available for source scan {source_scan_id}")
                        continue
                
                # Create analysis row
                analysis_row = {
                    'source_scan_id': source_scan_id,
                    'target_scan_id': target_row['scan_id'],
                    'source_store_planogram_id': source_row['store_planogram_id'],
                    'target_store_planogram_id': target_row['store_planogram_id'],
                    'source_planogram_name': source_row['planogram_name'],
                    'target_planogram_name': target_row['planogram_name'],
                    'source_section_id': source_row['section_id'],
                    'target_section_id': target_row['section_id'],
                    'source_section_name': source_row['section_name'],
                    'target_section_name': target_row['section_name'],
                    'source_is_additional_section': source_row['is_additional_section'],
                    'target_is_additional_section': target_row['is_additional_section'],
                    'source_pre_pog_percentage': source_row.get('pre_pog_percentage', 0),
                    'target_pre_pog_percentage': target_row.get('pre_pog_percentage', 0),
                    'source_post_pog_percentage': source_row.get('post_pog_percentage', 0),
                    'target_post_pog_percentage': target_row.get('post_pog_percentage', 0),
                    'source_pre_osa_percentage': source_row.get('pre_osa_percentage', 0),
                    'target_pre_osa_percentage': target_row.get('pre_osa_percentage', 0),
                    'source_post_osa_percentage': source_row.get('post_osa_percentage', 0),
                    'target_post_osa_percentage': target_row.get('post_osa_percentage', 0),
                    'source_ok_count': source_row.get('ok_count', 0),
                    'target_ok_count': target_row.get('ok_count', 0),
                    'source_wandering_count': source_row.get('wandering_count', 0),
                    'target_wandering_count': target_row.get('wandering_count', 0),
                    'source_oos_count': source_row.get('oos_count', 0),
                    'target_oos_count': target_row.get('oos_count', 0),
                    'source_hole_count': source_row.get('hole_count', 0),
                    'target_hole_count': target_row.get('hole_count', 0),
                    'comment': ''
                }
                
                # Add comments based on differences
                source_pre_pog = source_row.get('pre_pog_percentage', 0) or 0
                target_pre_pog = target_row.get('pre_pog_percentage', 0) or 0
                source_is_additional = source_row.get('is_additional_section', False)
                target_is_additional = target_row.get('is_additional_section', False)
                
                # Determine comment based on priority order
                if source_row['planogram_name'] != target_row['planogram_name']:
                    analysis_row['comment'] = 'Wrong POG Name Mapping'
                elif source_row['section_name'] != target_row['section_name']:
                    analysis_row['comment'] = 'Same POG Name but Different Section'
                elif not source_is_additional and target_is_additional:
                    analysis_row['comment'] = 'Target Has Additional Section (Source Does Not)'
                elif target_pre_pog > source_pre_pog:
                    analysis_row['comment'] = 'Target Has Higher POG% Than Source'
                else:
                    analysis_row['comment'] = 'No Issues'
                
                # Add MAv2_Map_by column with target map_by value only
                analysis_row['MAv2_Map_by'] = target_row.get('map_by_value', '')
                
                analysis_data.append(analysis_row)
            
            # Create analysis DataFrame
            analysis_df = pd.DataFrame(analysis_data)
            
            # Save analysis CSV
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analysis_with_comments_{timestamp}.csv"
            filepath = os.path.join(self.run_folder, filename)
            analysis_df.to_csv(filepath, index=False)
            
            # Create Excel with color highlighting
            excel_filename = self.create_excel_with_color_highlighting(analysis_df, filepath)
            
            return filepath
            
        except Exception as e:
            print(f"❌ Error creating analysis CSV with comments: {e}")
            return None

    def create_analysis_csv_with_comments(self, source_data, target_data):
        """Create analysis CSV with comments and highlighting"""
        try:
            # Create DataFrames
            source_df = pd.DataFrame(source_data)
            target_df = pd.DataFrame(target_data)
            
            # Create analysis data
            analysis_data = []
            
            # Use scan mapping to match source and target scans
            print(f"📊 Scan mapping available: {self.scan_mapping}")
            print(f"📊 Source scans: {source_df['scan_id'].tolist()}")
            print(f"📊 Target scans: {target_df['scan_id'].tolist()}")
            
            for _, source_row in source_df.iterrows():
                source_scan_id = source_row['scan_id']
                
                # Find corresponding target scan ID from mapping
                target_scan_id = None
                if self.scan_mapping and source_scan_id in self.scan_mapping:
                    target_scan_id = self.scan_mapping[source_scan_id]
                    print(f"🔍 Looking for target scan ID: {target_scan_id} (mapped from source {source_scan_id})")
                
                # Find target scan by target scan ID
                target_matches = target_df[target_df['scan_id'] == target_scan_id] if target_scan_id else pd.DataFrame()
                
                if len(target_matches) > 0:
                    target_row = target_matches.iloc[0]
                    print(f"✅ Matched source scan {source_scan_id} with target scan {target_row['scan_id']}")
                else:
                    # Fallback: try to match by index if no mapping found
                    if len(target_df) > 0:
                        target_row = target_df.iloc[0]  # Use first target scan as fallback
                        print(f"⚠️  No mapping found for source scan {source_scan_id}, using fallback target scan {target_row['scan_id']}")
                    else:
                        print(f"❌ No target data available for source scan {source_scan_id}")
                        continue
                
                # Create analysis row
                analysis_row = {
                        'source_scan_id': source_scan_id,
                        'target_scan_id': target_row['scan_id'],
                        'source_store_planogram_id': source_row['store_planogram_id'],
                        'target_store_planogram_id': target_row['store_planogram_id'],
                        'source_planogram_name': source_row['planogram_name'],
                        'target_planogram_name': target_row['planogram_name'],
                        'source_section_id': source_row['section_id'],
                        'target_section_id': target_row['section_id'],
                        'source_section_name': source_row['section_name'],
                        'target_section_name': target_row['section_name'],
                        'source_is_additional_section': source_row['is_additional_section'],
                        'target_is_additional_section': target_row['is_additional_section'],
                        'source_pre_pog_percentage': source_row.get('pre_compliance', 0),
                        'target_pre_pog_percentage': target_row.get('pre_compliance', 0),
                        'source_post_pog_percentage': source_row.get('post_compliance', 0),
                        'target_post_pog_percentage': target_row.get('post_compliance', 0),
                        'source_pre_osa_percentage': source_row.get('pre_osa', 0),
                        'target_pre_osa_percentage': target_row.get('pre_osa', 0),
                        'source_post_osa_percentage': source_row.get('post_osa', 0),
                        'target_post_osa_percentage': target_row.get('post_osa', 0),
                        'source_ok_count': source_row.get('ok_count', 0),
                        'target_ok_count': target_row.get('ok_count', 0),
                        'source_wandering_count': source_row.get('wandering_count', 0),
                        'target_wandering_count': target_row.get('wandering_count', 0),
                        'source_oos_count': source_row.get('oos_count', 0),
                        'target_oos_count': target_row.get('oos_count', 0),
                        'source_hole_count': source_row.get('hole_count', 0),
                        'target_hole_count': target_row.get('hole_count', 0),
                        'source_planogram_unique_count': source_row.get('planogram_unique_count', 0),
                        'target_planogram_unique_count': target_row.get('planogram_unique_count', 0),
                        'source_planogram_all_count': source_row.get('planogram_all_count', 0),
                        'target_planogram_all_count': target_row.get('planogram_all_count', 0),
                        'source_realogram_unique_count': source_row.get('realogram_unique_count', 0),
                        'target_realogram_unique_count': target_row.get('realogram_unique_count', 0),
                        'source_realogram_all_count': source_row.get('realogram_all_count', 0),
                        'target_realogram_all_count': target_row.get('realogram_all_count', 0),
                        'comment': ''
                    }
                    
                # Add comments based on differences
                source_pre_pog = source_row.get('pre_compliance', 0) or 0
                target_pre_pog = target_row.get('pre_compliance', 0) or 0
                
                # Use planogram name comparison instead of ID
                if source_row['planogram_name'] != target_row['planogram_name']:
                    analysis_row['comment'] = 'Different Store POG Mapped'
                elif source_row['section_name'] != target_row['section_name']:
                    analysis_row['comment'] = 'Different Section Mapped'
                elif target_pre_pog > source_pre_pog:
                    analysis_row['comment'] = 'Better Mapping (Higher Target POG)'
                else:
                    analysis_row['comment'] = 'No Issues'
                
                # Add MAv2_Map_by column with target map_by value only
                analysis_row['MAv2_Map_by'] = target_row.get('map_by_value', '')
                
                analysis_data.append(analysis_row)
            
            # Create analysis DataFrame
            analysis_df = pd.DataFrame(analysis_data)
            
            # Save analysis CSV
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analysis_with_comments_{timestamp}.csv"
            filepath = os.path.join(self.run_folder, filename)
            analysis_df.to_csv(filepath, index=False)
            
            # Create Excel with color highlighting
            excel_filename = self.create_excel_with_color_highlighting(analysis_df, filepath)
            
            return filepath
            
        except Exception as e:
            print(f"❌ Error creating analysis CSV with comments: {e}")
            return None
    
    def create_excel_with_color_highlighting(self, df, csv_filename):
        """Create Excel file with color highlighting based on comments"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            
            # Generate Excel filename
            excel_filename = csv_filename.replace('.csv', '_highlighted.xlsx')
            
            # Create Excel writer
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Analysis', index=False)
                
                # Get the worksheet
                worksheet = writer.sheets['Analysis']
                
                # Define color schemes
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')      # Red for wrong POG name mapping
                orange_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')  # Orange for same POG but different section
                yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Yellow for target additional section
                blue_fill = PatternFill(start_color='CCE6FF', end_color='CCE6FF', fill_type='solid')     # Blue for higher target POG%
                green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')   # Green for no issues
                
                # Apply highlighting based on comments
                for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
                    comment = df.iloc[row-2]['comment']
                    
                    if comment == 'Wrong POG Name Mapping':
                        # Highlight entire row in red
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = red_fill
                    elif comment == 'Same POG Name but Different Section':
                        # Highlight entire row in orange
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = orange_fill
                    elif comment == 'Target Has Additional Section (Source Does Not)':
                        # Highlight entire row in yellow
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = yellow_fill
                    elif comment == 'Target Has Higher POG% Than Source':
                        # Highlight entire row in blue
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = blue_fill
                    elif comment == 'No Issues':
                        # Highlight entire row in green
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = green_fill
            
            print(f"✅ Excel file with color highlighting created: {excel_filename}")
            return excel_filename
            
        except ImportError:
            print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            return None
    
    def run(self):
        """Run the complete SDK workflow"""
        print("🚀 SCAN COPY SDK - Starting Workflow")
        print("=" * 60)
        
        # Step 1: Configuration
        if not self.step1_configuration():
            return False
        
        # Step 1.5: Get results path
        if not self.step1_5_get_results_path():
            return False
        
        # Create results folder for this run (after path selection)
        if not self.create_results_folder():
            return False
        print()
        
        # Step 2: Get source scan IDs
        if not self.step2_get_source_scan_ids():
            return False
        
        # Step 3: Get target store
        if not self.step3_get_target_store():
            return False
        
        # Step 4: Run copy script
        if not self.step4_run_copy_script():
            return False
        
        # Step 5: Get target scan IDs (auto-extract from mapping if available)
        if not self.step5_get_target_scan_ids():
            return False
        
        # Generate analysis CSV
        if not self.generate_analysis_csv():
            return False
        
        print("\n" + "=" * 60)
        print("🎉 WORKFLOW COMPLETED SUCCESSFULLY!")
        print("=" * 60)
        print("Generated files:")
        print("- Source scan analysis CSV")
        print("- Target scan analysis CSV (if target scan IDs provided)")
        print("- Combined scan mapping CSV")
        
        return True

def main():
    """Main function"""
    sdk = ScanCopySDK()
    success = sdk.run()
    
    if not success:
        print("\n❌ Workflow failed. Please check the errors above.")
        sys.exit(1)
    else:
        print("\n✅ All steps completed successfully!")

if __name__ == "__main__":
    main()
